"""
WhatsApp Blasting Views
Handles customer group management and blast messaging campaigns
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_http_methods
from django.db import transaction
from django.utils import timezone as dj_timezone
from django.contrib import messages
import openpyxl
import csv
import io
import json
import time
import logging

from .models import (
    Tenant, TenantUser, Customer, WhatsAppConnection,
    CustomerGroup, GroupMember, BlastCampaign, BlastRecipient,
    Contest, ContestEntry, Conversation, CoreMessage
)
from .whatsapp_service import WhatsAppAPIService
from .blast_tasks import send_blast_campaign_task

logger = logging.getLogger(__name__)

# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def get_tenant_from_request(request):
    """Get tenant from logged in user"""
    try:
        tenant_user = TenantUser.objects.select_related('tenant').get(user=request.user)
        return tenant_user.tenant
    except TenantUser.DoesNotExist:
        return None

def get_whatsapp_connection(tenant):
    """Get the first WhatsApp connection for tenant"""
    return WhatsAppConnection.objects.filter(tenant=tenant).first()

# =============================================================================
# CUSTOMER GROUP VIEWS
# =============================================================================

@login_required
def blast_groups_list(request):
    """List all customer groups"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return redirect('auth_login')
    
    groups = CustomerGroup.objects.filter(tenant=tenant).prefetch_related('members')
    contests = Contest.objects.filter(tenant=tenant, is_active=True)
    
    context = {
        'tenant': tenant,
        'groups': groups,
        'contests': contests,
        'user': request.user,
    }
    return render(request, 'messaging/blast_groups_list.html', context)

@login_required
def blast_group_detail(request, group_id):
    """View details of a specific customer group"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return redirect('auth_login')
    
    group = get_object_or_404(CustomerGroup, group_id=group_id, tenant=tenant)
    members = GroupMember.objects.filter(group=group).select_related('customer')
    
    context = {
        'tenant': tenant,
        'group': group,
        'members': members,
        'user': request.user,
    }
    return render(request, 'messaging/blast_group_detail.html', context)

@login_required
@require_http_methods(["POST"])
def blast_create_group(request):
    """Create a new customer group"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        name = request.POST.get('name')
        description = request.POST.get('description', '')
        
        if not name:
            return JsonResponse({'success': False, 'error': 'Group name is required'}, status=400)
        
        group = CustomerGroup.objects.create(
            tenant=tenant,
            name=name,
            description=description,
            source='manual'
        )
        
        messages.success(request, f'Group "{name}" created successfully!')
        return JsonResponse({
            'success': True,
            'group_id': str(group.group_id),
            'redirect_url': f'/blast/groups/{group.group_id}/'
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def blast_import_group(request):
    """Import customer group from Excel/CSV file"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        group_name = request.POST.get('group_name')
        file = request.FILES.get('file')
        
        if not group_name or not file:
            return JsonResponse({'success': False, 'error': 'Group name and file are required'}, status=400)
        
        # Determine file type
        file_ext = file.name.split('.')[-1].lower()
        
        customers_data = []
        
        if file_ext in ['xlsx', 'xls']:
            # Handle Excel files
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            # Get header row
            headers = [cell.value for cell in ws[1]]
            
            # Expected columns: name, phone_number, and optional fields
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                customer_data = {}
                for idx, header in enumerate(headers):
                    if idx < len(row) and row[idx]:
                        customer_data[header.lower().replace(' ', '_')] = row[idx]
                
                customers_data.append(customer_data)
        
        elif file_ext == 'csv':
            # Handle CSV files
            file_content = file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(file_content))
            
            for row in csv_reader:
                customer_data = {k.lower().replace(' ', '_'): v for k, v in row.items() if v}
                customers_data.append(customer_data)
        
        else:
            return JsonResponse({'success': False, 'error': 'Unsupported file format. Use .xlsx, .xls, or .csv'}, status=400)
        
        # Create group and customers
        with transaction.atomic():
            group = CustomerGroup.objects.create(
                tenant=tenant,
                name=group_name,
                description=f'Imported from {file.name}',
                source='import',
                import_file_name=file.name,
                import_date=dj_timezone.now()
            )
            
            created_count = 0
            existing_count = 0
            
            for data in customers_data:
                # Validate required fields
                if 'name' not in data or 'phone_number' not in data:
                    continue
                
                # Clean phone number
                phone = str(data['phone_number']).strip()
                if not phone.startswith('+'):
                    phone = f"+{phone}"
                
                # Get or create customer
                customer, created = Customer.objects.get_or_create(
                    tenant=tenant,
                    phone_number=phone,
                    defaults={
                        'name': data.get('name', ''),
                        'gender': data.get('gender', 'N/A'),
                        'age': data.get('age') if 'age' in data else None,
                        'city': data.get('city'),
                        'state': data.get('state'),
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    existing_count += 1
                
                # Add to group
                GroupMember.objects.get_or_create(
                    tenant=tenant,
                    group=group,
                    customer=customer
                )
            
            messages.success(
                request,
                f'Group "{group_name}" created with {created_count} new customers and {existing_count} existing customers!'
            )
            
            return JsonResponse({
                'success': True,
                'group_id': str(group.group_id),
                'created_count': created_count,
                'existing_count': existing_count,
                'total_count': created_count + existing_count
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def blast_create_from_contest(request):
    """Create a customer group from contest participants"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        contest_id = request.POST.get('contest_id')
        group_name = request.POST.get('group_name')
        include_verified_only = request.POST.get('verified_only', 'false') == 'true'
        
        if not contest_id or not group_name:
            return JsonResponse({'success': False, 'error': 'Contest and group name are required'}, status=400)
        
        contest = get_object_or_404(Contest, contest_id=contest_id, tenant=tenant)
        
        # Get contest entries
        entries = ContestEntry.objects.filter(contest=contest).select_related('customer')
        
        if include_verified_only:
            entries = entries.filter(is_verified=True)
        
        with transaction.atomic():
            # Create group
            group = CustomerGroup.objects.create(
                tenant=tenant,
                name=group_name,
                description=f'Participants from contest: {contest.name}',
                source='contest',
                contest=contest
            )
            
            # Add participants to group
            for entry in entries:
                GroupMember.objects.get_or_create(
                    tenant=tenant,
                    group=group,
                    customer=entry.customer
                )
            
            member_count = group.members.count()
            messages.success(request, f'Group "{group_name}" created with {member_count} participants!')
            
            return JsonResponse({
                'success': True,
                'group_id': str(group.group_id),
                'member_count': member_count
            })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def blast_delete_group(request, group_id):
    """Delete a customer group"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        group = get_object_or_404(CustomerGroup, group_id=group_id, tenant=tenant)
        group_name = group.name
        group.delete()
        
        messages.success(request, f'Group "{group_name}" deleted successfully!')
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

# =============================================================================
# BLAST CAMPAIGN VIEWS
# =============================================================================

@login_required
def blast_campaigns_list(request):
    """List all blast campaigns"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return redirect('auth_login')
    
    campaigns = BlastCampaign.objects.filter(tenant=tenant).prefetch_related('target_groups', 'target_contests')
    
    context = {
        'tenant': tenant,
        'campaigns': campaigns,
        'user': request.user,
    }
    return render(request, 'messaging/blast_campaigns_list.html', context)

@login_required
def blast_create_campaign(request):
    """Create a new blast campaign"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return redirect('auth_login')
    
    groups = CustomerGroup.objects.filter(tenant=tenant)
    contests = Contest.objects.filter(tenant=tenant)
    whatsapp_connection = get_whatsapp_connection(tenant)
    
    if request.method == 'POST':
        try:
            name = request.POST.get('name')
            message_text = request.POST.get('message_text')
            message_image_url = request.POST.get('message_image_url', '')
            
            selected_groups = request.POST.getlist('target_groups')
            selected_contests = request.POST.getlist('target_contests')
            
            if not name or not message_text:
                messages.error(request, 'Campaign name and message are required!')
                return redirect('blast_create_campaign')
            
            if not whatsapp_connection:
                messages.error(request, 'No WhatsApp connection found. Please set up WhatsApp first.')
                return redirect('blast_create_campaign')
            
            with transaction.atomic():
                # Create blast campaign
                campaign = BlastCampaign.objects.create(
                    tenant=tenant,
                    whatsapp_connection=whatsapp_connection,
                    name=name,
                    message_text=message_text,
                    message_image_url=message_image_url if message_image_url else None,
                    status='draft'
                )
                
                # Add target groups
                if selected_groups:
                    campaign.target_groups.set(selected_groups)
                
                # Add target contests
                if selected_contests:
                    campaign.target_contests.set(selected_contests)
                
                # Calculate total recipients
                recipients = set()
                
                # Add group members
                for group_id in selected_groups:
                    group = CustomerGroup.objects.get(group_id=group_id)
                    for member in group.members.all():
                        recipients.add(member.customer)
                
                # Add contest participants
                for contest_id in selected_contests:
                    contest = Contest.objects.get(contest_id=contest_id)
                    for entry in contest.entries.all():
                        recipients.add(entry.customer)
                
                # Create blast recipients
                for customer in recipients:
                    BlastRecipient.objects.create(
                        tenant=tenant,
                        blast_campaign=campaign,
                        customer=customer,
                        status='pending'
                    )
                
                campaign.total_recipients = len(recipients)
                campaign.save()
                
                messages.success(request, f'Campaign "{name}" created with {len(recipients)} recipients!')
                return redirect('blast_campaign_detail', blast_id=campaign.blast_id)
        
        except Exception as e:
            messages.error(request, f'Error creating campaign: {str(e)}')
            return redirect('blast_create_campaign')
    
    context = {
        'tenant': tenant,
        'groups': groups,
        'contests': contests,
        'user': request.user,
    }
    return render(request, 'messaging/blast_create_campaign.html', context)

@login_required
def blast_campaign_detail(request, blast_id):
    """View details of a blast campaign"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return redirect('auth_login')
    
    campaign = get_object_or_404(BlastCampaign, blast_id=blast_id, tenant=tenant)
    recipients = BlastRecipient.objects.filter(blast_campaign=campaign).select_related('customer')
    
    context = {
        'tenant': tenant,
        'campaign': campaign,
        'recipients': recipients,
        'user': request.user,
    }
    return render(request, 'messaging/blast_campaign_detail.html', context)

@login_required
@require_http_methods(["POST"])
def blast_send_campaign(request, blast_id):
    """Send a blast campaign via WABot WhatsApp API"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        campaign = get_object_or_404(BlastCampaign, blast_id=blast_id, tenant=tenant)
        
        if campaign.status not in ['draft', 'failed']:
            return JsonResponse({'success': False, 'error': 'Campaign cannot be sent in current status'}, status=400)
        
        # Update campaign status
        campaign.status = 'sending'
        campaign.started_at = dj_timezone.now()
        campaign.save()
        
        # Check if we should send async (for large campaigns)
        use_async = request.POST.get('async', 'false') == 'true' or campaign.total_recipients > 50
        
        if use_async:
            # For large campaigns, we would use Celery here
            # For now, we'll use threading to not block the request
            import threading
            thread = threading.Thread(
                target=send_blast_campaign_task,
                args=(campaign.blast_id,)
            )
            thread.daemon = True
            thread.start()
            
            messages.success(request, f'Campaign "{campaign.name}" is being sent in the background. You can check the progress on the campaign detail page.')
            return JsonResponse({
                'success': True,
                'message': 'Campaign is being sent in the background',
                'async': True,
                'total_recipients': campaign.total_recipients
            })
        else:
            # Send synchronously for small campaigns
            result = send_blast_campaign_task(campaign.blast_id)
            
            # Prepare response message
            if result['success']:
                sent_count = result['sent_count']
                failed_count = result['failed_count']
                
                if sent_count > 0 and failed_count > 0:
                    msg = f'Campaign sent with partial success: {sent_count} sent, {failed_count} failed'
                elif sent_count > 0:
                    msg = f'Campaign sent successfully to {sent_count} recipients!'
                else:
                    msg = f'Campaign failed: {failed_count} messages could not be sent'
                
                messages.success(request, msg)
                
                return JsonResponse({
                    'success': True,
                    'message': msg,
                    'sent_count': sent_count,
                    'failed_count': failed_count,
                    'total_recipients': campaign.total_recipients
                })
            else:
                messages.error(request, f'Campaign failed: {result.get("error")}')
                return JsonResponse({'success': False, 'error': result.get('error')}, status=500)
    
    except Exception as e:
        logger.error(f"Error in blast_send_campaign: {str(e)}")
        # Update campaign status to failed
        try:
            campaign.status = 'failed'
            campaign.save()
        except:
            pass
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
@require_http_methods(["POST"])
def blast_cancel_campaign(request, blast_id):
    """Cancel a blast campaign"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        campaign = get_object_or_404(BlastCampaign, blast_id=blast_id, tenant=tenant)
        
        if campaign.status not in ['draft', 'scheduled', 'sending']:
            return JsonResponse({'success': False, 'error': 'Campaign cannot be cancelled'}, status=400)
        
        campaign.status = 'cancelled'
        campaign.save()
        
        messages.success(request, f'Campaign "{campaign.name}" has been cancelled!')
        return JsonResponse({'success': True})
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

@login_required
def blast_campaign_progress(request, blast_id):
    """Get real-time progress of a blast campaign"""
    tenant = get_tenant_from_request(request)
    if not tenant:
        return JsonResponse({'success': False, 'error': 'Not authenticated'}, status=401)
    
    try:
        campaign = get_object_or_404(BlastCampaign, blast_id=blast_id, tenant=tenant)
        
        # Count recipients by status
        status_counts = {}
        recipients = BlastRecipient.objects.filter(blast_campaign=campaign)
        
        for status in ['pending', 'queued', 'sent', 'delivered', 'failed', 'skipped']:
            status_counts[status] = recipients.filter(status=status).count()
        
        return JsonResponse({
            'success': True,
            'campaign_status': campaign.status,
            'total_recipients': campaign.total_recipients,
            'sent_count': campaign.sent_count,
            'delivered_count': campaign.delivered_count,
            'failed_count': campaign.failed_count,
            'success_rate': campaign.success_rate,
            'status_counts': status_counts,
            'started_at': campaign.started_at.isoformat() if campaign.started_at else None,
            'completed_at': campaign.completed_at.isoformat() if campaign.completed_at else None,
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

